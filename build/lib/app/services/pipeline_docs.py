from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

from app.models.schemas import DcpSpan
from app.services.orchestrator import Orchestrator


class DocumentPipeline:
    """
    Document pipeline (best-effort): extract text then run text detectors.

    Optional deps:
    - pdfplumber for PDF
    - python-docx for DOCX
    - openpyxl for XLSX
    """

    def __init__(self, orchestrator: Orchestrator | None = None) -> None:
        self.orc = orchestrator or Orchestrator()

    def extract_text(self, file_path: str) -> str:
        p = Path(file_path)
        suffix = p.suffix.lower()

        if suffix in {".txt", ".md", ".log"}:
            return p.read_text(encoding="utf-8", errors="ignore")

        if suffix == ".pdf":
            try:
                import pdfplumber
            except Exception as e:
                raise RuntimeError("pdfplumber non installé. Fais: pip install pdfplumber") from e

            parts: List[str] = []
            with pdfplumber.open(str(p)) as pdf:
                for page in pdf.pages:
                    parts.append(page.extract_text() or "")
            return "\n".join(parts)

        if suffix == ".docx":
            try:
                import docx
            except Exception as e:
                raise RuntimeError("python-docx non installé. Fais: pip install python-docx") from e
            d = docx.Document(str(p))
            return "\n".join([para.text for para in d.paragraphs])

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except Exception as e:
                raise RuntimeError("openpyxl non installé. Fais: pip install openpyxl") from e

            wb = load_workbook(str(p), read_only=True, data_only=True)
            parts: List[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    vals = [str(v) for v in row if v is not None]
                    if vals:
                        parts.append(" ".join(vals))
            return "\n".join(parts)

        raise RuntimeError(f"Format non supporté: {suffix}")

    def detect_file(
        self,
        *,
        file_path: str,
        language: str = "fr",
        detectors: List[str] | None = None,
        min_score: float = 0.4,
        merge_overlaps: bool = True,
        return_text: bool = False,
    ) -> Tuple[List[DcpSpan], Dict[str, List[DcpSpan]], Dict[str, int], Dict[str, str]]:
        text = self.extract_text(file_path)
        detectors = detectors or ["regex", "presidio", "spacy", "hf"]
        return self.orc.detect_text_multi(
            text=text,
            language=language,
            detectors=detectors,
            min_score=min_score,
            merge_overlaps=merge_overlaps,
            return_text=return_text,
            best_effort=True,
        )