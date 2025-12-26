from __future__ import annotations

from app.extractors.base import BaseExtractor, ExtractResult, normalize_path


class XlsxExtractor:
    name = "xlsx"
    supported_suffixes = {".xlsx"}

    def can_extract(self, path: str) -> bool:
        p = normalize_path(path)
        return p.is_file() and p.suffix.lower() in self.supported_suffixes

    def extract(self, path: str) -> ExtractResult:
        p = normalize_path(path)
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl") from e

        wb = load_workbook(str(p), read_only=True, data_only=True)
        parts: list[str] = []
        cell_count = 0

        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                vals = [str(v) for v in row if v is not None and str(v).strip()]
                if vals:
                    parts.append(" ".join(vals))
                    cell_count += len(vals)

        return ExtractResult(
            text="\n".join(parts).strip(),
            source_path=str(p),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata={"extractor": self.name, "sheets": str(len(wb.worksheets)), "cells": str(cell_count)},
        )